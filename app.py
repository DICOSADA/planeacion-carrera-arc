import os
import shutil
from datetime import datetime, timedelta
import streamlit as st
from openpyxl import load_workbook
import pandas as pd

# --- CONFIGURACIÓN DE LA PÁGINA WEB ---
st.set_page_config(
    page_title="Planeación Carrera ARC",
    page_icon="⚓",
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- CONTROL DE ACCESO SEGURO ---
password = st.text_input("Ingresa la contraseña de acceso:", type="password")

if password != st.secrets["PASSWORD_SECRETA"]:
    st.warning("Por favor, ingresa la contraseña para ver la aplicación.")
    st.stop() # <--- Esto es lo que detiene la app para que no se vea nada más

# --- RUTAS Y VARIABLES GLOBALES (Modificado para Descargas en Celular Android) ---
ARCHIVO_EXCEL = "PROYECCION OFICIALES.xlsx"

OPCIONES_COMPLEMENTACION = [
    "1. ADM. MARITIMA",
    "2. ADMINISTRACION CIM",
    "3. ELECTRONICA",
    "4. INGENIERIA",
    "5. OCEANOGRAFIA",
    "6. SUP. ADMINISTRACION"
]

TIEMPOS_GRADOS_1 = {"TK": 4, "TF": 4, "TN": 5, "CC": 5, "CF": 5, "CN": 5}
TIEMPOS_GRADOS_2 = {"ST": 4, "TE": 4, "CT": 5, "MY": 5, "TC": 5, "CR": 5}
SECUENCIA_1 = ["TK", "TF", "TN", "CC", "CF", "CN"]
SECUENCIA_2 = ["ST", "TE", "CT", "MY", "TC", "CR"]

def calcular_proximos_ascensos(grado_actual, fecha_ascenso_str):
    if not grado_actual or not fecha_ascenso_str:
        return []
    g_upper = grado_actual.strip().upper()
    secuencia = SECUENCIA_1 if g_upper in SECUENCIA_1 else (SECUENCIA_2 if g_upper in SECUENCIA_2 else None)
    tiempos = TIEMPOS_GRADOS_1 if g_upper in SECUENCIA_1 else (TIEMPOS_GRADOS_2 if g_upper in SECUENCIA_2 else None)
    
    if not secuencia: return []
    try:
        fecha_base = datetime.strptime(fecha_ascenso_str.split(" ")[0], "%Y-%m-%d")
        idx_actual = secuencia.index(g_upper)
    except ValueError:
        return []
        
    proximos = []
    fecha_acumulada = fecha_base
    for i in range(idx_actual + 1, len(secuencia)):
        grado_siguiente = secuencia[i]
        anos_grado_anterior = tiempos[secuencia[i-1]]
        try:
            nuevo_anio = fecha_acumulada.year + anos_grado_anterior
            fecha_acumulada = fecha_acumulada.replace(year=nuevo_anio)
        except ValueError:
            fecha_acumulada = fecha_acumulada + timedelta(days=anos_grado_anterior * 365)
        proximos.append((grado_siguiente, fecha_acumulada.strftime("%Y-%m-%d")))
    return proximos

@st.cache_data(ttl=600)
def cargar_base_datos_web():
    if not os.path.exists(ARCHIVO_EXCEL):
        return None, f"❌ Archivo Excel no encontrado en la ruta: {ARCHIVO_EXCEL}. Asegúrate de descargarlo en la carpeta Descargas de tu celular."
    
    try:
        wb = load_workbook(ARCHIVO_EXCEL, data_only=True)
        hoja = wb.active
        base_oficiales = {}
        anios_disponibles = set()
        unidades_disponibles = set()
        grados, siglas, especialidades, unidades = set(), set(), set(), set()
        
        mapeo_cursos = {"CM": "Comando", "VC": "Virtual Comando", "CEM": "Curso de Estado Mayor", "VB": "Virtual Basico"}
        mapeo_estados = {
            "EMB6": "Falta 6 meses embarque", "EMB1": "Falta 1 año embarque",
            "EMB11/2": "Falta 1 año y medio de embarque", "EMB2": "Falta 2 años de embarque",
            "MANDO1": "Falta 1 año de mando", "MANDO6": "Falta 6 meses de mando", "RETIRO": "Pendiente por retirar"
        }
        
        columnas_proyeccion = {}
        ultimo_ano = "2026"
        for col_idx in range(11, 73):
            ano_celda = str(hoja.cell(row=7, column=col_idx).value or "").strip()
            if ano_celda and ano_celda.isdigit(): ultimo_ano = ano_celda
            anios_disponibles.add(ultimo_ano)
            semestre = "1 Semestre" if col_idx % 2 != 0 else "2 Semestre"
            columnas_proyeccion[col_idx] = (semestre, ultimo_ano)

        for fila in range(8, hoja.max_row + 1):
            cedula = str(hoja.cell(row=fila, column=5).value or "").strip()
            if not cedula or cedula == "CEDULA": continue
            
            ant_cruda = str(hoja.cell(row=fila, column=1).value or "").strip()
            try: antiguedad = int(float(ant_cruda))
            except ValueError: antiguedad = 999999
            
            fecha_cruda = str(hoja.cell(row=fila, column=10).value or "").strip()
            fecha_limpia = fecha_cruda.split(" ")[0] if " " in fecha_cruda else fecha_cruda
            
            g = str(hoja.cell(row=fila, column=3).value or "").strip()
            s = str(hoja.cell(row=fila, column=4).value or "").strip()
            u = str(hoja.cell(row=fila, column=8).value or "").strip()
            dep = str(hoja.cell(row=fila, column=9).value or "").strip()
            nombres = str(hoja.cell(row=fila, column=6).value or "").strip()
            
            celda_g = hoja.cell(row=fila, column=7)
            e = str(celda_g.value or "").strip()
            
            ya_complemento = False
            if celda_g.font and celda_g.font.color and celda_g.font.color.rgb:
                if "00b050" in str(celda_g.font.color.rgb).lower():
                    ya_complemento = True
            
            estado_complementacion = " (Ya complemento)" if ya_complemento else " (No ha complementado)"
            proyeccion_individual, cursos_por_periodo = [], set()
            
            for col_idx, (semestre, ano) in columnas_proyeccion.items():
                valor_celda = str(hoja.cell(row=fila, column=col_idx).value or "").strip()
                if valor_celda and valor_celda != "None" and valor_celda != "✓":
                    if valor_celda == "CEM":
                        texto_final = f"Curso de Estado Mayor ({ano})"
                        cursos_por_periodo.add((valor_celda, semestre, ano))
                    elif valor_celda == "COMPLE":
                        texto_final = f"Proyectado complementación ({ano})"
                        cursos_por_periodo.add((valor_celda, semestre, ano))
                    elif valor_celda in mapeo_cursos:
                        texto_final = f"{mapeo_cursos[valor_celda]} ({semestre} {ano})"
                        cursos_por_periodo.add((valor_celda, semestre, ano))
                    elif valor_celda in ["1", "2", "3", "4"]:
                        texto_final = f"Basico ({valor_celda} trimestre {ano})"
                        cursos_por_periodo.add((valor_celda, semestre, ano))
                    elif valor_celda in mapeo_estados:
                        texto_final = f"{mapeo_estados[valor_celda]}"
                    else:
                        texto_final = f"{valor_celda} ({semestre} {ano})"
                    proyeccion_individual.append(texto_final)
            
            if g: grados.add(g)
            if s: siglas.add(s)
            if e: especialidades.add(e)
            if u: 
                unidades.add(u)
                unidades_disponibles.add(u)
                
            base_oficiales[cedula] = {
                "cedula": cedula, "antiguedad": antiguedad, "grado": g, "esp_sigla": s, "nombres": nombres,
                "especialidad": e, "unidad": u, "depende": dep, "ascenso": fecha_limpia,
                "estado_comple": estado_complementacion, "ya_complemento": ya_complemento,
                "proyeccion": proyeccion_individual, "cursos_periodo": cursos_por_periodo
            }
        return base_oficiales, anios_disponibles, unidades_disponibles, sorted(list(grados)), sorted(list(siglas)), sorted(list(especialidades)), sorted(list(unidades))
    except Exception as e:
        return None, f"❌ Error: {str(e)}"

# --- INTERFAZ WEB STREAMLIT ---
st.title("⚓ Sistema de Planeación y Gestión de Carrera ARC")

datos = cargar_base_datos_web()
if datos[0] is None:
    st.error(datos[1])
    st.stop()

base_oficiales, anios_disponibles, unidades_disponibles, lista_grados, lista_siglas, lista_esp, lista_uni = datos

menu = st.sidebar.selectbox("Navegación", ["🔍 Buscador Individual", "📋 Filtros Masivos", "📊 Estadísticas", "⚓ Embarque y Mando"])

if menu == "🔍 Buscador Individual":
    st.subheader("Buscador Individual de Oficial")
    criterio = st.text_input("Ingrese Cédula o Apellidos:").strip()
    
    if criterio:
        oficial = None
        if criterio in base_oficiales:
            oficial = base_oficiales[criterio]
        else:
            for ced, ofi in base_oficiales.items():
                if criterio.upper() in ofi['nombres'].upper():
                    oficial = ofi
                    break
        if oficial:
            col1, col2 = st.columns(2)
            with col1:
                st.markdown(f"**Antigüedad:** {oficial['antiguedad'] if oficial['antiguedad'] != 999999 else '---'}")
                st.markdown(f"**Nombres:** {oficial['nombres']}")
                st.markdown(f"**Grado y Sigla:** {oficial['grado']} - {oficial['esp_sigla']}")
                st.markdown(f"**Especialidad:** {oficial['especialidad']}{oficial['estado_comple']}")
                st.markdown(f"**Unidad:** {oficial['unidad']} ({oficial['depende']})")
                st.markdown(f"**Último Ascenso:** {oficial['ascenso']}")
            with col2:
                st.markdown("### 🚀 Proyección Semestral")
                for item in oficial['proyeccion']:
                    st.markdown(f"- {item}")
                proximos = calcular_proximos_ascensos(oficial['grado'], oficial['ascenso'])
                if proximos:
                    st.markdown("### 🎖️ Próximos Ascensos")
                    for g_prox, f_prox in proximos:
                        st.info(f"{g_prox} - Fecha estimada: {f_prox}")
        else:
            st.warning("❌ No se encontró ningún oficial con ese criterio.")

elif menu == "📋 Filtros Masivos":
    st.subheader("Filtros Masivos de Oficiales")
    c1, c2, c3 = st.columns(3)
    with c1:
        f_grado = st.selectbox("Grado", ["TODOS"] + lista_grados)
    with c2:
        f_sigla = st.selectbox("Sigla", ["TODOS"] + lista_siglas)
    with c3:
        f_unidad = st.selectbox("Unidad", ["TODOS"] + lista_uni)
        
    resultados = []
    for ced, ofi in base_oficiales.items():
        if f_grado != "TODOS" and ofi['grado'] != f_grado: continue
        if f_sigla != "TODOS" and ofi['esp_sigla'] != f_sigla: continue
        if f_unidad != "TODOS" and ofi['unidad'] != f_unidad: continue
        resultados.append(ofi)
        
    st.write(f"**Total encontrados:** {len(resultados)}")
    if resultados:
        df_res = pd.DataFrame([{
            "Antigüedad": o['antiguedad'] if o['antiguedad'] != 999999 else "---",
            "Cédula": o['cedula'],
            "Grado": o['grado'],
            "Sigla": o['esp_sigla'],
            "Nombres": o['nombres'],
            "Especialidad": o['especialidad'],
            "Unidad": o['unidad']
        } for o in resultados])
        st.dataframe(df_res, use_container_width=True)

elif menu == "📊 Estadísticas":
    st.subheader("📊 Panel Estadístico")
    tipo_stat = st.selectbox("Indicador", ["Cursos Básicos", "Cursos de Comando", "Carga por Unidad (Top 10)"])
    
    conteo = {}
    for ofi in base_oficiales.values():
        for cod, sem, ano_curso in ofi['cursos_periodo']:
            if tipo_stat == "Cursos Básicos" and cod in ["1", "2", "3", "4", "VB"]:
                conteo[ano_curso] = conteo.get(ano_curso, 0) + 1
            elif tipo_stat == "Cursos de Comando" and cod in ["CM", "VC"]:
                conteo[ano_curso] = conteo.get(ano_curso, 0) + 1
            elif tipo_stat == "Carga por Unidad (Top 10)":
                uni = ofi['unidad'] or "SIN UNIDAD"
                conteo[uni] = conteo.get(uni, 0) + 1

    if conteo:
        df_stat = pd.DataFrame(list(conteo.items()), columns=["Categoría", "Cantidad"]).sort_values(by="Cantidad", ascending=False)
        if "Top" in tipo_stat: df_stat = df_stat.head(10)
        st.bar_chart(df_stat.set_index("Categoría"))
    else:
        st.warning("No hay datos suficientes para graficar.")

elif menu == "⚓ Embarque y Mando":
    st.subheader("Oficiales con Requerimientos de Embarque y Mando")
    condiciones_validas = {"EMB6", "EMB1", "EMB11/2", "EMB2", "MANDO6", "MANDO1", "MANDO11/2", "MANDO2"}
    
    tripulantes = []
    wb = load_workbook(ARCHIVO_EXCEL, data_only=True)
    hoja = wb.active
    for fila in range(8, hoja.max_row + 1):
        cedula = str(hoja.cell(row=fila, column=5).value or "").strip()
        if not cedula or cedula == "CEDULA": continue
        for col_idx in range(11, 73):
            val_celda = str(hoja.cell(row=fila, column=col_idx).value or "").strip().replace(" ", "")
            if any(cond in val_celda for cond in condiciones_validas):
                tripulantes.append({
                    "Grado": hoja.cell(row=fila, column=3).value,
                    "Sigla": hoja.cell(row=fila, column=4).value,
                    "Cédula": cedula,
                    "Nombres": hoja.cell(row=fila, column=6).value,
                    "Unidad": hoja.cell(row=fila, column=8).value,
                    "Condición": val_celda
                })
                break
    if tripulantes:
        st.dataframe(pd.DataFrame(tripulantes), use_container_width=True)
    else:
        st.success("No hay oficiales pendientes por embarque o mando.")

st.sidebar.markdown("---")
st.sidebar.markdown("by S1 PEÑA DIEGO")
